import time
import random
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select

# 크롬 드라이버 자동 업데이트
from webdriver_manager.chrome import ChromeDriverManager

# 크롬 옵션
chrome_options = Options()
user_data = r"C:\Users\zofld\AppData\Local\Google\Chrome\User Data"
chrome_options.add_argument(f"user-data-dir={user_data}")
chrome_options.add_experimental_option("detach", True)

service = Service(executable_path=ChromeDriverManager().install())
driver = webdriver.Chrome(service=service, options=chrome_options)

# 상점 열기
driver.implicitly_wait(3) # 3초 동안 기다림
# driver.maximize_window() # 화면 최대화함
driver.get('https://store.gaijin.net/user.php?skin_lang=en')

# 로그인 버튼 클릭
try:
    if driver.find_element(By.CSS_SELECTOR, "#__container > div.row.js-account-list-container > div:nth-child(1) > ul > li > a").is_displayed():
        print("로그인 버튼 존재")
        btn_login = driver.find_element(By.CSS_SELECTOR, "#__container > div.row.js-account-list-container > div:nth-child(1) > ul > li > a")
        btn_login.click()
        time.sleep(random.uniform(1, 2))
except:
    print("로그인 버튼 없음")

# 가이진 코인 충전 페이지 이동
print("충전 페이지 이동")
btn_balance = driver.find_element(By.CSS_SELECTOR, "#gaijin-menu > nav > ul > li:nth-child(4) > a")
btn_balance.click()
time.sleep(random.uniform(1, 2))

# 1코인 선택
print("1코인 선택")
btn_coin = driver.find_element(By.CSS_SELECTOR, "#bodyRoot > div.content > section > div.balance > div.balance__wrapper > div.balance__refill.balance-refill > ul > li:nth-child(1) > div > div.balance-refill__fix-sum-link.js-fix-refill.js-popup__trigger")
btn_coin.click()
time.sleep(random.uniform(1, 2))

# 결제창 클릭
print("결제창 클릭")
btn_confirm = driver.find_element(By.CSS_SELECTOR, "#g2s_cc_usd")
btn_confirm.click()
driver.implicitly_wait(4)


# 결제창 프레임 전환
print("결제창 대기")
driver.switch_to.frame(driver.find_element(By.XPATH, '//*[@id="pay-frame"]'))
driver.implicitly_wait(3)


path = r'F:\workspace\gaijin\data\data.xlsx'
wb = openpyxl.load_workbook(path)
ws = wb.active # 현재 활성화 된 시트 선택

select = Select(driver.find_element(By.XPATH, '//*[@id="currency_lbl"]/div/select'))

row = 2
for item in select.options:
    get_value = item.get_attribute('innerText')
    get_character = item.get_attribute('value')
    select.select_by_value(get_value)
    time.sleep(1)
    get_currency = driver.find_element(By.CSS_SELECTOR, "#orderTotalAmount")
    currency = get_currency.text # $ 1.00
    arr = currency.split(maxsplit=1) # 1.00
    change_currency = arr[1].strip()
    print(get_character + ' ' + change_currency)
    ws[f'A{row}'] = str(get_character)
    ws[f'B{row}'] = str(change_currency)
    row = row + 1

print("출력 종료")

# 종료
wb.save(path)
driver.close()